# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render,get_object_or_404,redirect,render_to_response
from django.shortcuts import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from . import forms
from datetime import datetime, timedelta
from tablib import Dataset
import openpyxl
from openpyxl import workbook
# from BallbearingSite.libraries import PyJdate
from django.conf import settings
from BallbearingSite.forms import UserForm,UserProfileInfoForm,User,UploadFileForm,ExportSpec,SendEmailForm,ContactUsForm,StocksForm
from BallbearingSite.models import UserProfileInfo,Posts,ExpireDays,SellerDesktop,Cars,Parts
from django.contrib.auth import authenticate,login,logout
from django.http import HttpResponseRedirect, HttpResponse
from django.core.urlresolvers import reverse
from django.contrib.auth.decorators import login_required
from django.utils.translation import ugettext as _
from import_export.formats.base_formats import XLS
from django.core.mail import send_mail
from django.http import HttpResponseRedirect
from django.template import RequestContext
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.cache import cache_page
from django.http import JsonResponse
from dal import autocomplete
# import requests
import json
# Create your views here.


from django.views.generic.edit import FormView

global expiredays,duration
expiredays=ExpireDays.objects.filter(id=1)[0].days
duration = datetime.today() - timedelta(expiredays)

@login_required
def sendemailview(request):
    form=SendEmailForm(None)
    if request.method == "POST":
        if request.POST.get("expire&mail"):
            form=SendEmailForm(data=request.POST)
            if form.is_valid():
                subject=form.cleaned_data.get("subject")
                mailmessage=form.cleaned_data.get("message")
                recievers=Stocks.objects.filter(date__lte=duration,confirm=_('approved')).values_list('user__email', flat=True).distinct() #lte = less than date & gte =greater than
                # send_mail(subject,mailmessage,"golshan.zahra@gmail.com",["golshan.zahra@gmail.com","golshan_zahra500@yahoo.com",],fail_silently=False)
                messages.success(request,"   با موفقیت ارسال گردید"  )
                q=Stocks.objects.all().filter(date__lte=duration).update(confirm=_('expired')) #update is to set a field to a particular value for all the objects in a QuerySet
        if request.POST.get("sendmailtoall"):
            form=SendEmailForm(data=request.POST)
            if form.is_valid():
                subject=form.cleaned_data.get("subject")
                mailmessage=form.cleaned_data.get("message")
                recievers=Stocks.objects.filter().values_list('user__email', flat=True).distinct()
                # send_mail(subject,mailmessage,"golshan.zahra@gmail.com",["golshan.zahra@gmail.com","golshan_zahra500@yahoo.com",],fail_silently=False)
                # print(recievers)
                messages.success(request,"   با موفقیت ارسال گردید"  )
        if request.POST.get("expire"):
                form=SendEmailForm(data=request.POST)
                q=Stocks.objects.all().filter(date__lte=duration).update(confirm=_('expired')) #update is to set a field to a particular value for all the objects in a QuerySet
                messages.success(request," کالاها با موفقیت منقضی گردیدند"  )


    context = {
        "mailform":form,
    }
    return render(request,'BallbearingSite/controlsite.html',context)

def homepage(request):
    # return HttpResponse(settings.LOCALE_PATHS)

    return render(request,'BallbearingSite/homepage.html')

def index(request):

    stocks_list=Parts.objects.all().filter(confirm=_('approved')).order_by('-date')
    posts=Posts.objects.all().order_by('-createDateTime')[:3]
    companyname=UserProfileInfo.objects.values('city').distinct()
    cars=Cars.objects.filter(confirm=_('approved')).order_by('carname')

    if request.GET.get('search'):
             stocks_list =Parts.objects.all().filter(confirm=_('approved')).order_by('-date')
             car=request.GET.get('dropdowncars')
             city=request.GET.get('dropdowncity')
             partname=request.GET.get('partname')
             # if city is not None :
             #      stocks_list = stocks_list.filter(city__icontains = city)
             if request.GET.get("dropdowncars")!=" همه خودروها":
                  stocks_list = stocks_list.filter(car__carname__exact = car)
             if request.GET.get("dropdowncity")!="همه شهرها":
                  stocks_list = stocks_list.filter(user__in=(UserProfileInfo.objects.filter(city=city).values_list("user")))
             if request.GET.get("partname") is not None :
                  stocks_list = stocks_list.filter(partname__icontains = partname)


    page = request.GET.get('page', 1)
    paginator = Paginator(stocks_list, 5)
    try:
        st = paginator.page(page)
    except PageNotAnInteger:
        st = paginator.page(1)
    except EmptyPage:
        st = paginator.page(paginator.num_pages)

    context={
        'allstocks':stocks_list,
        'posts': posts,
        'companyname':companyname,
        'cars':cars,
        'st':st
    }
    return render(request,'BallbearingSite/index.html',context)


def news(request):
    posts=Posts.objects.all()
    context={
    'posts': posts,
    }
    return render(request,'BallbearingSite/news.html',context)


@login_required
def special(request):
    return HttpResponse("You are logged in, Nice!")


@login_required
def user_logout(request):
    logout(request)
    return HttpResponseRedirect(reverse('BallbearingSite:index'))


def register(request):
    registered = False
    user_form =UserForm(None)
    profile_form = UserProfileInfoForm(None)

    if request.method == "POST":
        user_form =UserForm(data=request.POST)
        profile_form = UserProfileInfoForm(data=request.POST)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            user.set_password(user.password) #hashing password
            user.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()
            messages.success(request,"با موفقیت ثبت گردید"  )
        else:
            messages.error(request, "خطا !")


    else:
        user_form=UserForm()
        profile_form = UserProfileInfoForm()
    return render(request,'BallbearingSite/register.html',
                           {'user_form':user_form,
                            'profile_form':profile_form,
                               'registered':registered})


def user_login(request):
    if request.method=='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')

        user=authenticate(username=username,password=password)

        if user:
            if user.is_active:
                login(request,user)
                return redirect(reverse('BallbearingSite:index'))

            else:
                return HttpResponse("ACCOUNT NOT ACTIVE")
        else:
            messages.error(request, "نام کاربری یا رمزعبور اشتباه است !")

    return render(request,'BallbearingSite/login.html')


def stock(request):
    st_form=StocksForm(None)
    # # TODO:admin have to complete his profile and if is_seller=="" returns error
    # if request.user.is_authenticated():
    #     is_seller=UserProfileInfo.objects.filter(user=request.user).values("is_seller")[0]['is_seller']
    # else:
    #     is_seller=""
    if request.method == "POST":
        st_form =StocksForm(request.POST, request.FILES) #request.FILES is related to pic
        if st_form.is_valid():
            instance=st_form.save()
            # id_name=(st_form.cleaned_data.get("name"))
            # instance.name=StocksName.objects.filter(id=id_name).values("name")[0]['name'] #becouse of autocomplete it was saving the value by its id and i converted to name
            instance.user=request.user
            # instance.pic=st_form.cleaned_data.get['pic']
            instance.save()
            messages.success(request,"با موفقیت ثبت گردید" ,extra_tags="savestock")

        else:
            messages.error(request, "خطا !")

    else:
        st_form=StocksForm()
    # return render(request,'BallbearingSite/stock.html',{'stocks_form':st_form,'seller':is_seller})
    return render(request,'BallbearingSite/stock.html',{'stocks_form':st_form,})


def mystocks_view(request):
    if request.method=='POST': #if anything to delete
        todel = request.POST.getlist('todelete')
        Parts.objects.filter(user=request.user,id__in=todel).delete()
    stocks_list=Parts.objects.all().filter(user__username=request.user).order_by('-date') #The username which is forign key of user :user__username
     #paginating for table
    page = request.GET.get('page', 1)
    paginator = Paginator(stocks_list, 15)
    try:
        myst = paginator.page(page)
        # print(stocks_list.count())
    except PageNotAnInteger:
        myst = paginator.page(1)
    except EmptyPage:
        myst = paginator.page(paginator.num_pages)

    return render(request,'BallbearingSite/mystocks.html',{'myst':myst,'mystocks':stocks_list,})


def allstocks_view(request,username=None):
    # stocks_list=Parts.objects.all().filter(confirm=_('approved') ).order_by('-date')

    # if request.GET.get('search'):
    #      stocks_list =Stocks.objects.all().filter(confirm=_('approved') )
    #      name=request.GET.get('namesearch')
    #      number=request.GET.get('numbersearch')
    #      brand=request.GET.get('brandsearch')
    #      city=request.GET.get('citysearch')
    #      companyname=request.GET.get('companynamesearch')
    #      if name is not None :
    #           stocks_list = stocks_list.filter(name__icontains = name)
    #      if number is not None :
    #           stocks_list = stocks_list.filter(number__icontains = number)
    #      if request.GET.get("brandsearch"):
    #           stocks_list = stocks_list.filter(brand__icontains = brand)
    #      if request.GET.get("citysearch"):
    #           stocks_list = stocks_list.filter(user__in=(UserProfileInfo.objects.filter(city=city).values_list("user")))
    #      if request.GET.get("companynamesearch"):
    #           stocks_list = stocks_list.filter(user__in=(UserProfileInfo.objects.filter(companyname__icontains=companyname).values_list("user")))
         # print(question_set.filter(user__in=(UserProfileInfo.objects.filter(city=city).values_list("user")))     )
         # print(stocks_list)


    # if request.method=='POST': #if has POST to pass to desktop
    #    if request.POST.get("toseller"):#if send to seller button clicked
    #       tosave = request.POST.getlist('sendtoseller')
    #       for i in  range(len(tosave)):
    #          userseller=Stocks.objects.filter(id=tosave[i]).values_list('user',flat=True)[0]
    #          SellerDesktop.objects.create(buyer=request.user,stock=Stocks.objects.get(id=tosave[i]),seller=User.objects.get(id=userseller))
    #       messages.success(request," با موفقیت برای فروشنده ارسال گردید" )

    stocks_list=Parts.objects.all().filter(confirm=_('approved'),user=User.objects.filter(username=username).values_list("id") ).order_by('-date')
          #paginating for table
    page = request.GET.get('page', 1)
    paginator = Paginator(stocks_list, 15)
    try:
        myst = paginator.page(page)
    except PageNotAnInteger:
        myst = paginator.page(1)
    except EmptyPage:
        myst = paginator.page(paginator.num_pages)

    context={
            'allstocks':stocks_list,
            'myst':myst,

        }
    return render(request,'BallbearingSite/sellerstocks.html',context)


def buyerstocks_view(request):
    stocks_list=Stocks.objects.all().filter(confirm=_('approved') ).order_by('-date')

      #paginating for table
    page = request.GET.get('page', 1)
    paginator = Paginator(stocks_list, 15)
    try:
        myst = paginator.page(page)
    except PageNotAnInteger:
        myst = paginator.page(1)
    except EmptyPage:
        myst = paginator.page(paginator.num_pages)

    context={
        'allstocks':stocks_list,
        'myst':myst,
    }
    return render(request,'BallbearingSite/buyerstocks.html',context)


# def stock_delete(request, id=None):
#     deleted=False
#     instance=get_object_or_404(Stocks,id=id)
#     instance.delete()
#     deleted=True
#     messages.success(request,"با موفقیت حذف گردید" ,extra_tags="deletestock")
#     stocks_list=Stocks.objects.all().filter(user__username=request.user) #The username which is forign key of user :user__username
#     stocks_dict={'mystocks':stocks_list}
#     return HttpResponseRedirect(instance.get_absolute_url())
#     return render(request,'BallbearingSite/mystocks.html',context=stocks_dict)


# def stock_bulk_delete(request, id=None):
#     # if request.POST.get('delete'):
#       # Stocks.objects.filter(id__in=request.POST.getlist('items')).delete()
#     # if request.method=='POST':
#     todel = request.POST.getlist('todelete')
#     Stocks.objects.filter(user=request.user,id__in=todel).delete()
#
#     messages.success(request,"با موفقیت حذف گردید" )
#     stocks_list=Stocks.objects.all().filter(user__username=request.user) #The username which is forign key of user :user__username
#     stocks_dict={'mystocks':stocks_list}
#     return HttpResponseRedirect(instance.get_absolute_url())
#     return render(request,'BallbearingSite/mystocks.html',context=stocks_dict)


def stock_update(request, id=None):
    # instance=Stocks.objects.get(id=id)
    #Filling the form with old informations
    instance = get_object_or_404(Parts, id=id)
    stocks_form =StocksForm(instance=instance)
    # save new informations
    if request.method == "POST":
        stocks_form=StocksForm(request.POST, request.FILES)
        if stocks_form.is_valid():
              cd = stocks_form.cleaned_data
              # old=Stocks.objects.get(id=id)
              old=Parts.objects.get(id=id)
              old.price=cd.get('price')
              old.car=cd.get('car')
              old.partname=cd.get('partname')
              old.number=cd.get('number')
              old.description=cd.get('description')
              old.pic=cd.get('pic')
              print(cd.get('pic'))
              old.price=cd.get('price')
              # old.seller=cd.get('seller')
              old.date=datetime.today()
              old.confirm=_('pending')
              old.save()
              # messages.success(request,_('Saved Successfully'),extra_tags="savestock")
              return HttpResponseRedirect('/mystocks')
    return render(request,'BallbearingSite/stock.html',{'stocks_form':stocks_form})




    # def ca_import(request):
#     uploadform=UploadFileForm(None)
#     if request.method == 'POST' :
#         uploadform = UploadFileForm(request.POST, request.FILES)
#         if uploadform.is_valid():
#             file = uploadform.cleaned_data['docfile']
#             data = bytes()
#             for chunk in file.chunks():
#                 data += chunk
#             dataset = XLS().create_dataset(data)
#
#             user_ids = []
#             user=request.user
#             user_ids.append(user.id)
#             dataset.insert_col(0, user_ids, 'user')
#
#
#
#             result = ExportSpec().import_data(dataset,dry_run=False, raise_errors=True, )
#     return render(request, 'BallbearingSite/news.html',{'uploadform':uploadform})


#__PDF INSERT__
def ca_import(request):
    uploadform=UploadFileForm(None)
    if request.method == 'POST' :
        uploadform = UploadFileForm(request.POST, request.FILES)
        if uploadform.is_valid():
            file = uploadform.cleaned_data['docfile']
            workbook = openpyxl.load_workbook(filename=file, read_only=True)
            # Get name of the first sheet and then open sheet by name
            first_sheet = workbook.get_sheet_names()[0]
            worksheet = workbook.get_sheet_by_name(first_sheet)
            data = []
            # data2 = []

            try:
            #  if uploadform.cleaned_data.get('seller') == 'True':
            #     sel=True
            #  else:
            #     sel=False
             for row in worksheet.iter_rows(row_offset=1): # Offset for header
                # stock = Stocks()
                part = Parts()
                part.user = request.user
                # stock.name = row[0].value
                Cars.objects.get_or_create(carname=str( row[0].value)) #Save new car in Car table if not available (for autocomplete)
                part.car = Cars.objects.get(carname=str( row[0].value)) #car is a foreign key filed related to Car table
                # stock.number = row[1].value
                part.partname = row[1].value
                # stock.suffix = row[2].value
                # stock.pic = row[2].value
                # stock.brand = row[3].value.upper()
                part.description = row[2].value
                # stock.comment = row[4].value
                part.price = row[3].value
                # stock.date = datetime.today()
                # stock.confirm = 'approved'
                # stock.seller = sel
                data.append(part)
            #Bulk create data
             # Stocks.objects.bulk_create(data)
             Parts.objects.bulk_create(data)

            #save for autocomplete
             # for row in range(2,worksheet.max_row+1):
             #    for column in "A":  #Here you can add or reduce the columns
             #        cell_name = "{}{}".format(column, row)
             #        Cars.objects.get_or_create(carname=str(worksheet[cell_name].value))


             messages.success(request,"با موفقیت ثبت گردید" ,extra_tags="saveexcel")
            except :
             messages.error(request,_('نکات ذیل را برای ثبت فایل درنظر بگیرید. از خالی نبود فیلدهای name و brand اطمینان حاصل کنید'),extra_tags="excelerror")

    return render(request, 'BallbearingSite/excelfile.html',{'uploadform':uploadform})

def detailnewsview(request,id=None):
    postdetail=get_object_or_404(Posts,id=id)
    context={
    'postdetail':postdetail,
    }
    return render(request, 'BallbearingSite/detailnews.html',context)

def detailadvertisementsview(request,id=None):
    advertisementdetail=get_object_or_404(Parts,id=id)

    if request.method=='POST': #if has POST to pass to desktop
       if request.POST.get("toseller"):#if send to seller button clicked
             userseller=Parts.objects.filter(id=id).values_list('user',flat=True)[0]
             print(userseller)
             SellerDesktop.objects.create(buyer=request.user,stock=Parts.objects.get(id=id),seller=User.objects.get(id=userseller),date=datetime.today())
       messages.success(request," با موفقیت برای فروشنده ارسال گردید" )

    context={
    'advertisement':advertisementdetail,

    }
    return render(request, 'BallbearingSite/detailadvertisement.html',context)


def ContactUsView(request):
    contact_form=StocksForm(None)
    if request.method == "POST":
        contact_form =ContactUsForm(data=request.POST)
        if contact_form.is_valid():
            contact_form.save()

            messages.success(request,"با موفقیت ثبت گردید" ,extra_tags="savestock")
        else:
            messages.error(request, "خطا !")

    else:
        contact_form=ContactUsForm()
    return render(request,'BallbearingSite/contactus.html',{'contact_form':contact_form})


# AUTOCOMPLETE
class StocksAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        if not self.request.user.is_authenticated():
            return Cars.objects.none()
        qs = Cars.objects.filter(confirm=_('approved')).distinct()

        if self.q:
            qs = qs.filter(carname__istartswith=self.q,confirm=_('approved'))
        return qs
    def has_add_permission(self, request): #The user should have add permision inorder to be able to add new item ***important
        return True

    # def create_object(self, text):
    #       # print('aaaaaaaaa')
    #       # print(self. request.POST.get())
    #       # cd=StocksForm(request.POST).cleaned_data
    #       # print (cd.get('number'))
    #       return self.get_queryset().create(**{self.create_field: text, 'user' : self.request.user , 'number' : '111' , })

    # def post(self, request):
    #     """Create an object given a text after checking permissions."""
    #     # if not self.has_add_permission(request):
    #     #     return http.HttpResponseForbidden()
    #
    #     # if not self.create_field:
    #     #     raise ImproperlyConfigured('Missing "create_field"')
    #
    #     text = request.POST.get('text', None)
    #
    #     result = self.create_object(text)
    #
    #     return http.JsonResponse({
    #         'id': result.pk,
    #         'text': self.get_result_label(result),
    #     })

def mydesktop_view(request):
    if request.method=='POST': #if anything to delete
        todel = request.POST.getlist('todelete')
        SellerDesktop.objects.filter(seller=request.user,id__in=todel).delete()
    mylist=SellerDesktop.objects.filter(seller=request.user)
     #paginating for table
    page = request.GET.get('page', 1)
    paginator = Paginator(mylist, 15)
    try:
        myst = paginator.page(page)
        # print(stocks_list.count())
    except PageNotAnInteger:
        myst = paginator.page(1)
    except EmptyPage:
        myst = paginator.page(paginator.num_pages)

    return render(request,'BallbearingSite/mydesktop.html',{'myst':myst,'mydesktoplist':mylist,})

# def search(request):
#     if request.method=='POST':
#       city = request.POST.get('dropdowncity')
#       cars=Cars.objects.all()
#       return render(request, 'BallbearingSite/search.html',{'city':city,'cars':cars})

# TODO: this search is suspended untill insert completed
              # stocks_list=Stocks.objects.all().filter(confirm=_('approved') ).order_by('-date')
     # if request.GET.get('search'):
     #               stocks_list =Stocks.objects.all().filter(confirm=_('approved') )
     #               description=request.GET.get('desctiptionsearch')
     #               car=request.GET.get('carsearch')
     #               city=request.GET.get('citysearch')
     #               if car is not None :
     #                    stocks_list = stocks_list.filter(name__icontains = name)
     #               if number is not None :
     #                    stocks_list = stocks_list.filter(number__icontains = number)
     #               if request.GET.get("brandsearch"):
     #                    stocks_list = stocks_list.filter(brand__icontains = brand)
     #               if request.GET.get("citysearch"):
     #                    stocks_list = stocks_list.filter(user__in=(UserProfileInfo.objects.filter(city=city).values_list("user")))
     #               if request.GET.get("companynamesearch"):
     #                    stocks_list = stocks_list.filter(user__in=(UserProfileInfo.objects.filter(companyname__icontains=companyname).values_list("user")))
